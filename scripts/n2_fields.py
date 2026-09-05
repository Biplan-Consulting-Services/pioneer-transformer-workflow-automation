# -*- coding: utf-8 -*-
"""Build the CAML field definitions for N2.

Using createfieldasxml rather than the UI means the internal Name is set
explicitly in the same call as the DisplayName -- no create-short-then-rename
dance, and no 32-character truncation surprise, because we choose the internal
name outright.

Type mapping is deliberate, not a copy of the source:
  DateTime   -> DateOnly   copying DateTime reintroduces the UTC-midnight bug
  Choice     -> Text       a synced Choice rejects out-of-domain values silently
  Lookup     -> Text       a lookup ID means nothing across lists; sync the label
  MultiChoice-> Note       exports as a JSON array, can be long
"""
import openpyxl, json, io, re

WA = r"C:\Users\solei\OneDrive\Documents\Biplan\claude\Clients\Pioneer Transformer\Workflow-Automation"
PICKER = WA + r"\sharepoint-lists\Order Items parent column picker 2026-09-05.xlsx"

DROP = {"Order - Client", "Model - Client", "Mod. Rev. - Client",
        "Client - Client_ID", "Order - SA", "Order - Lead Time"}
# these two overwrite existing lookups instead of being created
RENAME_NOT_CREATE = {"Order - Model", "Order - Model Revision"}

PREFIX = {"Order": "Ord", "Models": "Mdl", "Model Revisions": "Rev", "Clients": "Cli"}

ws = openpyxl.load_workbook(PICKER, data_only=True)["Columns"]
hdr = [str(c.value).strip() for c in ws[1]]
ix = {h: i for i, h in enumerate(hdr)}

def caml_type(src):
    return {"Text": ("Text", ""), "Note": ("Note", ' NumLines="4"'),
            "Number": ("Number", ""), "Currency": ("Currency", ""),
            "Boolean": ("Boolean", ""), "URL": ("URL", ' Format="Hyperlink"'),
            "DateTime": ("DateTime", ' Format="DateOnly"'),
            "Choice": ("Text", ""), "MultiChoice": ("Note", ' NumLines="4"'),
            "Lookup": ("Text", "")}.get(src, ("Text", ""))

fields, seen = [], set()
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r[0]:
        continue
    row = {h: (str(r[i]).strip() if r[i] is not None else "") for h, i in ix.items()}
    if row["Include?"].lower() != "yes":
        continue
    disp = row["Proposed Order Items name"]
    if disp in DROP or disp in RENAME_NOT_CREATE:
        continue
    src = row["Type"]
    ctype, extra = caml_type(src)
    stem = re.sub(r"[^A-Za-z0-9]", "", disp.split(" - ", 1)[-1])
    name = (PREFIX.get(row["List"], "X") + stem)[:32]
    n, base = 2, name
    while name.lower() in seen:
        name = (base[:30] + str(n)); n += 1
    seen.add(name.lower())
    fields.append({
        "display": disp, "name": name, "type": ctype, "srcType": src,
        "list": row["List"], "srcCol": row["Display name"],
        "xml": '<Field Type="%s" DisplayName="%s" Name="%s" StaticName="%s" Required="FALSE" Group="Parent Sync"%s />'
               % (ctype, disp.replace("&", "&amp;"), name, name, extra)})

io.open("n2_fields.json", "w", encoding="utf-8").write(json.dumps(fields, indent=1))
print("fields to create: %d\n" % len(fields))
print("%-32s %-24s %-10s %s" % ("DISPLAY", "INTERNAL", "CREATE AS", "(source)"))
for f in fields:
    flag = "  <-- remapped" if f["type"] != f["srcType"] else ""
    print("%-32s %-24s %-10s %s%s" % (f["display"][:32], f["name"], f["type"], f["srcType"], flag))
print("\nlongest internal name: %d chars" % max(len(f["name"]) for f in fields))
print("all unique:", len({f["name"].lower() for f in fields}) == len(fields))
