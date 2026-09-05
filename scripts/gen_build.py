# -*- coding: utf-8 -*-
"""Turn the returned picker into an ordered create-list for N2, applying the
decisions taken after the picker was filled in -- which override some of its
own rows, so those are called out rather than silently dropped."""
import openpyxl, io, csv, collections

WA = r"C:\Users\solei\OneDrive\Documents\Biplan\claude\Clients\Pioneer Transformer\Workflow-Automation"
PICKER = WA + r"\sharepoint-lists\Order Items parent column picker 2026-09-05.xlsx"
OUT_MD = WA + r"\docs\n2-column-build-sheet.md"
OUT_CSV = WA + r"\sharepoint-lists\N2 column build list 2026-09-05.csv"

wb = openpyxl.load_workbook(PICKER, data_only=True)
ws = wb["Columns"]
hdr = [str(c.value).strip() if c.value else "" for c in ws[1]]
ix = {h: i for i, h in enumerate(hdr)}
rows = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r[0]:
        continue
    rows.append({h: (str(r[i]).strip() if r[i] is not None else "") for h, i in ix.items()})

yes = [r for r in rows if r["Include?"].lower() == "yes"]

# Decisions taken after the picker was filled in. Each overrides a Yes row.
OVERRIDE = {
    "Order - Client":        ("DROP",   "Client stays client — the existing Order Items.Client lookup is the only one."),
    "Model - Client":        ("DROP",   "Same decision. Reachable through the Client lookup."),
    "Mod. Rev. - Client":    ("DROP",   "Same decision."),
    "Client - Client_ID":    ("BACKFILL", "One-time backfill into the existing Client_ID_TextField, not a synced column."),
    "Order - Model":         ("RENAME", "NOT prefixed — overwrites the existing Order Items.Model lookup. SA units RE-RESOLVE, never copy."),
    "Order - Model Revision":("RENAME", "NOT prefixed — overwrites the existing Order Items.Model Revision lookup. Same SA rule."),
    "Order - SA":            ("DROP",   "Means 'this order should have SA units'. Beside SA Job it reads as a duplicate and is not."),
    "Order - Lead Time":     ("REPLACE", "Superseded: the client lead time comes from FRM13 via Clients.Lead Time, not Order.Lead Time (306 of 342 disagree)."),
    "Mod. Rev. - Duplicate Order": ("KEEP", "Create it, but it stays empty — 0 of 391 populated. Future engineering-completion logic fills it."),
    "Model - Latest Model Revision": ("KEEP", "Name it so it reads as 'the newest design', NOT 'this unit's revision'. They are different facts."),
}

buckets = collections.defaultdict(list)
for r in yes:
    name = r["Proposed Order Items name"]
    act, why = OVERRIDE.get(name, ("CREATE", ""))
    buckets[act].append((r, why))

with io.open(OUT_CSV, "w", encoding="utf-8-sig", newline="") as f:
    w = csv.writer(f)
    w.writerow(["Action", "Source list", "Source column", "Source internal name",
                "Type", "Proposed Order Items name", "Note"])
    for act in ("CREATE", "RENAME", "KEEP", "REPLACE", "BACKFILL", "DROP"):
        for r, why in sorted(buckets[act], key=lambda x: (x[0]["List"], x[0]["Display name"])):
            w.writerow([act, r["List"], r["Display name"], r["Internal name"],
                        r["Type"], r["Proposed Order Items name"], why or r["Notes"]])

n_create = len(buckets["CREATE"]) + len(buckets["KEEP"])
L = []; a = L.append
a("# N2 — the columns to create on `Order Items`")
a("")
a("Generated 2026-09-05 from the returned picker (`scripts/gen_build.py`, re-runnable).")
a("Full rows: **`sharepoint-lists/N2 column build list 2026-09-05.csv`**.")
a("")
a("The picker came back **57 Yes**. Ten of those were then overridden by decisions taken *after* it")
a("was filled in, so the real build is **%d new columns** — listed here rather than silently" % n_create)
a("dropped, because a future reader comparing the workbook to the list will otherwise think rows")
a("went missing.")
a("")
a("| Action | Count | Meaning |")
a("|---|---|---|")
for act, meaning in (("CREATE", "New prefixed column, straight from the picker."),
                     ("KEEP", "Create it, with a caveat attached."),
                     ("RENAME", "**Not** prefixed — overwrites an existing lookup."),
                     ("REPLACE", "Superseded by a different source."),
                     ("BACKFILL", "One-time fill of an existing column, not a synced column."),
                     ("DROP", "Decided against after the picker was returned.")):
    if buckets[act]:
        a("| **%s** | %d | %s |" % (act, len(buckets[act]), meaning))
a("")
a("## The ten overrides")
a("")
a("| Proposed name | Action | Why |")
a("|---|---|---|")
for act in ("RENAME", "REPLACE", "BACKFILL", "DROP", "KEEP"):
    for r, why in buckets[act]:
        if why:
            a("| `%s` | **%s** | %s |" % (r["Proposed Order Items name"], act, why))
a("")
a("## Creating them")
a("")
a("**Create each with a short name, then rename to the display name.** SharePoint derives the")
a("internal name from whatever the column is called at creation and never changes it afterwards, so")
a("creating `Order - Order Number` directly bakes in")
a("`Order_x0020__x002d__x0020_Order_`. Creating `OrderOrderNumber` and renaming gives a clean")
a("internal name and the display name you want.")
a("")
a("⚠️ Internal names are escaped **and truncated at 32 characters**. `Protector & Switchgear Item #`")
a("on this same list already became `Protector_x0020__x0026__x0020_Sw` — stopping mid-word. Anything")
a("writing an expression against a new column must read the internal name back from")
a("`_api/…/fields` (or an export's `ListSchema` record) rather than deriving it. See")
a("`a5-d1-d2-paste-sheet.md` for what that mistake costs.")
a("")
a("## By source list")
a("")
a("| Source | Columns to create |")
a("|---|---|")
per = collections.Counter(r["List"] for r, _ in buckets["CREATE"] + buckets["KEEP"])
for k, v in sorted(per.items()):
    a("| `%s` | %d |" % (k, v))
a("")
a("## Sequencing")
a("")
a("Creating the columns is safe on its own — an empty column changes nothing. **The sync flows")
a("(N3) wait for A3**, stripping 2c stage-stamping out of the trigger flow: before that, every")
a("synced write re-fires a ~100-action flow per row, which is the load shape that hit the capacity")
a("cap. Change-guard every write.")

io.open(OUT_MD, "w", encoding="utf-8").write("\n".join(L) + "\n")
print("picker rows %d | Yes %d" % (len(rows), len(yes)))
for act in ("CREATE", "KEEP", "RENAME", "REPLACE", "BACKFILL", "DROP"):
    if buckets[act]:
        print("  %-9s %d" % (act, len(buckets[act])))
print("real new columns:", n_create)
