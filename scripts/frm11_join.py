# -*- coding: utf-8 -*-
"""Reproduces the FRM11 <-> Order Items row-set table in
docs/frm11-coupling-analysis-2026-09-06.md.

FRM11's row set is not arbitrary: FournTank filters TableOrders on the unit ID,
drops anything already tanked, and dedups. This script re-applies those rules to
the SharePoint export and reports what each one accounts for, so a future
divergence is attributable rather than mysterious.

Reads the live workbook's "fourn Tank" sheet (TableFournTank, header on row 4),
not the archive table.
"""
import os, sys, glob, collections
import openpyxl
from load_exports import load

WA = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WB = os.path.join(WA, "workbooks", "PRO1.FRM11 - Planification Approbation Cuve.xlsx")
KEY = "NUMÉRO DE CUVE"

# FournTank's own filter, verbatim from power-query/FRM11/FournTank.pq
EXCLUDE_SUBSTRINGS = ("SA", "W", "R", "OS")

# "Rows to purge" -> Already Tanked, from power-query/FRM11/Rows to purge.pq.
# The .pq works in FRM10-12's two-letter codes; the transfer flow converts those
# to display names on the way into SharePoint, so they are translated back here.
PURGE_LOCATIONS = {"Extérieur", "Test", "Finition", "Livraison"}   # XT TE FI LI
PURGE_TANKING_LOCATION = "Tanking"                                      # TA + Status ~ TE


def latest_order_items():
    hits = sorted(glob.glob(os.path.join(WA, "sharepoint-lists", "Order Items *.csv")))
    if not hits:
        raise SystemExit("no 'Order Items *.csv' export in sharepoint-lists/")
    return hits[-1]


def frm11_keys():
    wb = openpyxl.load_workbook(WB, data_only=True, read_only=True)
    ws = wb["fourn Tank"]
    header = [c.value for c in next(ws.iter_rows(min_row=4, max_row=4))]
    i = header.index(KEY)
    return set(str(r[i]).strip() for r in ws.iter_rows(min_row=5, values_only=True) if r[i])


def excluded(unit):
    return any(s in unit for s in EXCLUDE_SUBSTRINGS)


def already_tanked(row):
    loc = row.get("Location") or ""
    status = row.get("Status") or ""
    return loc in PURGE_LOCATIONS or (loc == PURGE_TANKING_LOCATION and "TE" in status)


def main():
    path = latest_order_items()
    oi = load(path)
    S = frm11_keys()
    T = set((r.get("Unit ID") or "").strip() for r in oi if (r.get("Unit ID") or "").strip())

    missing = [r for r in oi
               if (r.get("Unit ID") or "").strip()
               and r["Unit ID"].strip() not in S
               and not excluded(r["Unit ID"].strip())]
    purged = [r for r in missing if already_tanked(r)]
    left = [r for r in missing if not already_tanked(r)]
    false_pos = [r for r in oi if (r.get("Unit ID") or "").strip() in S and already_tanked(r)]

    print("Order Items export: %s" % os.path.basename(path))
    print("FRM11 workbook:     %s" % os.path.basename(WB))
    print()
    print("FRM11 live (TableFournTank)          %5d" % len(S))
    print("Order Items                          %5d" % len(T))
    print("matching on Unit ID                  %5d" % len(S & T))
    print("on Order Items, absent from FRM11    %5d" % len(T - S))
    print("  ... excluded by SA/W/R/OS filter   %5d" % len([t for t in (T - S) if excluded(t)]))
    print("  ... explained by Already-Tanked    %5d" % len(purged))
    print("  ... remaining                      %5d" % len(left))
    print("in FRM11, absent from Order Items    %5d" % len(S - T))
    print()
    print("purge-rule false positives (present in FRM11 yet Already-Tanked): %d" % len(false_pos))
    if false_pos:
        print("  ^ the rule no longer reproduces -- the vocabularies have moved apart")
    print()
    orders = collections.Counter(u.split("-")[0] for u in sorted(S - T))
    print("FRM11-only units, by order: %s" % dict(sorted(orders.items())))
    print()
    print("unexplained (%d): %s" % (len(left), ", ".join(sorted(r["Unit ID"] for r in left))))


if __name__ == "__main__":
    sys.exit(main())
